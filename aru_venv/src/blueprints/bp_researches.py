# import datetime
from datetime import datetime
import io
import pandas as pd
from db.engine import get_session
from flask import Blueprint, jsonify, request, make_response, render_template_string
from models.Author import Author
from models.AuthorResearch import AuthorResearch
from models.Campus import Campus
from models.Departments import Departments
from models.ResearchDepartment import ResearchDepartment
from models.ResearchKeywords import ResearchKeywords
from models.ResearchInstAgenda import ResearchInstAgenda
from models.InstAgenda import InstAgenda
from models.ResearchDeptAgenda import ResearchDeptAgenda
from models.DeptAgenda import DeptAgenda
from sqlalchemy.orm import joinedload
from models.Research import Research
from sqlalchemy import or_, and_
import pdfkit
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font

bp = Blueprint("bp_researches", __name__, url_prefix="/v1/researches")


@bp.route("/main", methods=["GET", "POST"])
def handleResearchGetPost():
    session = get_session(autocommit=False, autoflush=False)()

    if request.method == "GET":
        try:
            search_query = request.args.get("q")
            startDate = request.args.get("startDate")
            endDate = request.args.get("endDate")
            presentation_location = request.args.get("presentation")
            publication_location = request.args.get("publication")
            department = request.args.get("department")
            camp_name = request.args.get("camp_name")

            query = (
                session.query(Research)
                .join(
                    ResearchDepartment,
                    Research.research_id == ResearchDepartment.research_id,
                )
                .join(Departments, ResearchDepartment.dept_id == Departments.dept_id)
                .join(Campus, Research.camp_id == Campus.camp_id)
                .outerjoin(
                    ResearchDeptAgenda,
                    Research.research_id == ResearchDeptAgenda.research_id,
                )
                .outerjoin(
                    DeptAgenda,
                    ResearchDeptAgenda.deptagenda_id == DeptAgenda.deptagenda_id,
                )
                .outerjoin(
                    ResearchInstAgenda,
                    Research.research_id == ResearchInstAgenda.research_id,
                )
                .outerjoin(
                    InstAgenda,
                    ResearchInstAgenda.instagenda_id == InstAgenda.instagenda_id,
                )
                .outerjoin(
                    AuthorResearch,
                    Research.research_id == AuthorResearch.research_id,
                )
                .outerjoin(
                    Author,
                    AuthorResearch.author_id == Author.author_id,
                )
                .options(
                    joinedload(Research.researchdepartment).joinedload(
                        ResearchDepartment.department
                    ),
                    joinedload(Research.authorresearch).joinedload(
                        AuthorResearch.author
                    ),
                    joinedload(Research.researchkeywords).joinedload(
                        ResearchKeywords.keywords
                    ),
                    joinedload(Research.researchdeptagenda).joinedload(
                        ResearchDeptAgenda.deptagenda
                    ),
                    joinedload(Research.researchinstagenda).joinedload(
                        ResearchInstAgenda.instagenda
                    ),
                )
            )

            # Check if all parameters are null
            if not any(
                [
                    search_query,
                    startDate,
                    endDate,
                    presentation_location,
                    publication_location,
                    department,
                    camp_name,
                ]
            ):
                # If all parameters are null, limit the results to 10
                query = query.limit(0)

            # Sidebar filters
            if department:
                departments_list = department.split(",")
                department_filters = [
                    Departments.dept_name.ilike(f"%{dept.strip()}%")
                    for dept in departments_list
                ]
                query = query.filter(or_(*department_filters))

            if search_query:
                search_query = f"%{search_query}%"
                query = query.filter(
                    or_(
                        Author.author_name.ilike(search_query),
                        Research.title.ilike(search_query),
                        Research.presented_where.ilike(search_query),
                        InstAgenda.instagenda_name.ilike(search_query),
                        DeptAgenda.deptagenda_name.ilike(search_query),
                        Research.abstract.ilike(search_query),
                    )
                )

            if presentation_location:
                presentation_list = presentation_location.split(",")
                presentation_filters = [
                    Research.presentation_location.ilike(f"%{location.strip()}%")
                    for location in presentation_list
                ]
                query = query.filter(or_(*presentation_filters))

            if publication_location:
                publication_list = publication_location.split(",")
                publication_filters = [
                    Research.published_where.ilike(f"%{location.strip()}%")
                    for location in publication_list
                ]
                query = query.filter(or_(*publication_filters))

                # Convert string dates to datetime objects if provided
            if startDate or endDate:
                if startDate:
                    startDate = datetime.strptime(startDate, "%Y-%m-%d")
                if endDate:
                    endDate = datetime.strptime(endDate, "%Y-%m-%d")

                # Apply filters based on the provided dates for cited_date, publication_date, and presentation_date
                if startDate and endDate:
                    query = query.filter(
                        or_(
                            and_(
                                Research.cited_date >= startDate,
                                Research.cited_date <= endDate,
                            ),
                            and_(
                                Research.publication_date >= startDate,
                                Research.publication_date <= endDate,
                            ),
                            and_(
                                Research.presentation_date >= startDate,
                                Research.presentation_date <= endDate,
                            ),
                        )
                    )
                elif startDate:
                    query = query.filter(
                        or_(
                            Research.cited_date >= startDate,
                            Research.publication_date >= startDate,
                            Research.presentation_date >= startDate,
                        )
                    )
                elif endDate:
                    query = query.filter(
                        or_(
                            Research.cited_date <= endDate,
                            Research.publication_date <= endDate,
                            Research.presentation_date <= endDate,
                        )
                    )

            if camp_name:
                camp_list = camp_name.split(",")
                camp_filters = [
                    Campus.camp_name.ilike(f"%{camp.strip()}%") for camp in camp_list
                ]
                query = query.filter(or_(*camp_filters))

            result = query.all()

            data = []
            for i in result:
                dept_names = [rd.department.dept_name for rd in i.researchdepartment]
                author_names = [ar.author.author_name for ar in i.authorresearch]
                keywords_names = [
                    rk.keywords.keywords_name for rk in i.researchkeywords
                ]
                dept_agendas = [
                    da.deptagenda.deptagenda_name for da in i.researchdeptagenda
                ]
                inst_agendas = [
                    ia.instagenda.instagenda_name for ia in i.researchinstagenda
                ]

                data.append(
                    {
                        "research_id": i.research_id,
                        "title": i.title,
                        "abstract": i.abstract,
                        "presented_where": i.presented_where,
                        "presentation_location": i.presentation_location,
                        "presentation_date": i.presentation_date,
                        "published_where": i.published_where,
                        "publication_date": i.publication_date,
                        "cited_where": i.cited_where,
                        "cited_date": i.cited_date,
                        "doi_or_full": i.doi_or_full,
                        "camp_name": i.campus.camp_name,
                        "user_id": i.user_id,
                        "dept_names": dept_names,
                        "author_names": author_names,
                        "keywords_names": keywords_names,
                        "dept_agendas": dept_agendas,
                        "inst_agendas": inst_agendas,
                    }
                )

            if data:
                return jsonify(
                    {"success": True, "message": "Fetched all Research", "data": data}
                ), 200
            else:
                return jsonify(
                    {"success": True, "message": "No research found", "data": []}
                ), 200
        except Exception as e:
            print(f"An error occurred: {e}")
            return jsonify({"success": False, "message": str(e)}), 500

    elif request.method == "POST":
        data = request.get_json()
        new_data = Research(
            title=data.get("title"),
            abstract=data.get("abstract"),
            presented_where=data.get("presented_where", None)
            and data["presented_where"],
            presentation_location=data.get("presentation_location", None)
            and data["presentation_location"],
            presentation_date=data.get("presentation_date", None)
            and data["presentation_date"],
            published_where=data.get("published_where", None)
            and data["published_where"],
            publication_date=data.get("publication_date", None)
            and data["publication_date"],
            cited_where=data.get("cited_where", None) and data["cited_where"],
            cited_date=data.get("cited_date", None) and data["cited_date"],
            doi_or_full=data.get("doi_or_full", None) and data["doi_or_full"],
            user_id=data.get("user_id"),
            camp_id=data.get("camp_id"),
        )
        try:
            session.add(new_data)
            session.commit()
            return jsonify(
                {
                    "success": True,
                    "message": str(new_data),
                    "research_id": new_data.research_id,
                }
            ), 201
        except Exception as e:
            session.rollback()
            print(f"An error occurred: {e}")
            return jsonify({"success": False, "message": str(e)}), 500
    else:
        return jsonify({"success": False, "message": "Invalid request"}), 400


@bp.route("/<int:id>", methods=["GET", "PUT", "DELETE"])
def handleResearchItemRequests(id):
    session = get_session(autocommit=False, autoflush=False)()
    if request.method == "GET":
        try:
            result = (
                session.query(Research)
                .options(
                    joinedload(Research.researchdepartment).joinedload("department"),
                    joinedload(Research.authorresearch).joinedload("author"),
                    joinedload(Research.researchkeywords).joinedload("keywords"),
                )
                .filter(Research.research_id == id)
                .first()
            )

            if result:
                dept_names = [
                    rd.department.dept_name
                    for rd in result.researchdepartment
                    if rd.department
                ]
                author_names = [
                    ar.author.author_name for ar in result.authorresearch if ar.author
                ]
                keyword_names = [
                    rk.keywords.keywords_name
                    for rk in result.researchkeywords
                    if rk.keywords
                ]

                return jsonify(
                    {
                        "success": True,
                        "message": "Fetched record",
                        "data": {
                            "research_id": result.research_id,
                            "title": result.title,
                            "abstract": result.abstract,
                            "presented_where": result.presented_where,
                            "presentation_location": result.presentation_location,
                            "presentation_date": result.presentation_date,
                            "published_where": result.published_where,
                            "publication_date": result.publication_date,
                            "cited_where": result.cited_where,
                            "cited_date": result.cited_date,
                            "doi_or_full": result.doi_or_full,
                            "camp_id": result.camp_id,
                            "user_id": result.user_id,
                            "dept_names": dept_names,
                            "author_names": author_names,
                            "keyword_names": keyword_names,
                        },
                    }
                ), 200
            else:
                return jsonify({"success": False, "message": "Research not found"}), 404
        except Exception as e:
            print(f"An error occurred: {e}")
            return jsonify({"success": False, "message": str(e)}), 500

    elif request.method == "PUT":
        try:
            result = session.query(Research).filter(Research.research_id == id).first()
            data = request.get_json()
            if result:
                result.title = data.get("title", result.title)
                result.abstract = data.get("abstract", result.abstract)
                result.presented_where = data.get(
                    "presented_where", result.presented_where
                )
                result.presentation_location = data.get(
                    "presentation_location", result.presentation_location
                )
                result.presentation_date = data.get(
                    "presentation_date", result.presentation_date
                )
                result.published_where = data.get(
                    "published_where", result.published_where
                )
                result.publication_date = data.get(
                    "publication_date", result.publication_date
                )
                result.cited_where = data.get("cited_where", result.cited_where)
                result.cited_date = data.get("cited_date", result.cited_date)
                result.doi_or_full = data.get("doi_or_full", result.doi_or_full)
                result.camp_id = data.get("camp_id", result.camp_id)
                result.user_id = data.get("user_id", result.user_id)

                session.commit()
                return jsonify({"success": True, "message": "Updated record."}), 200
            else:
                return jsonify({"success": False, "message": "Research not found"}), 404
        except Exception as e:
            session.rollback()
            print(f"An error occurred: {e}")
            return jsonify({"success": False, "message": str(e)}), 500

    elif request.method == "DELETE":
        try:
            result = session.query(Research).filter(Research.research_id == id).first()
            if not result:
                return jsonify({"success": False, "error": "Research not found"}), 404

            # Delete related records in the linking tables
            session.query(AuthorResearch).filter_by(research_id=id).delete()
            session.query(ResearchDepartment).filter_by(research_id=id).delete()
            session.query(ResearchKeywords).filter_by(research_id=id).delete()

            # Delete the research record
            session.delete(result)
            session.commit()

            return jsonify(
                {"success": True, "message": "Research deleted successfully"}
            ), 200
        except Exception as e:
            session.rollback()
            print(f"An error occurred: {e}")
            return jsonify({"success": False, "message": str(e)}), 500

    else:
        return jsonify({"success": False, "message": "Invalid request"}), 400


@bp.route("/data", methods=["GET"])
def getResearchData():
    session = get_session(autocommit=False, autoflush=False)()
    try:
        query = session.query(
            Research.research_id,
            Research.title,
            Research.abstract,
            Research.presented_where,
            Research.presentation_location,
            Research.presentation_date,
            Research.published_where,
            Research.publication_date,
            Research.cited_where,
            Research.cited_date,
            Research.doi_or_full,
            Research.camp_id,
            Research.user_id,
        )

        result = query.all()

        data = []
        for i in result:
            data.append(
                {
                    "research_id": i.research_id,
                    "title": i.title,
                    "abstract": i.abstract,
                    "presented_where": i.presented_where,
                    "presentation_location": i.presentation_location,
                    "presentation_date": i.presentation_date,
                    "published_where": i.published_where,
                    "publication_date": i.publication_date,
                    "cited_where": i.cited_where,
                    "cited_date": i.cited_date,
                    "doi_or_full": i.doi_or_full,
                    "camp_id": i.camp_id,
                    "user_id": i.user_id,
                }
            )

        return jsonify(
            {"success": True, "message": "Fetched Research Data", "data": data}
        ), 200
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@bp.route("/author", methods=["GET"])
# @jwt_required()
def handleAuthorResearches():
    session = get_session(autocommit=False, autoflush=False)()
    try:
        data = request.get_json()
        print(data.get("author_id"))
        results = (
            session.query(AuthorResearch)
            .filter(AuthorResearch.author_id == data.get("author_id"))
            .all()
        )

        if results:
            for result in results:
                res2 = session.query(AuthorResearch).all()
                authors = [i.author_id for i in res2]
                authors = [i.author_name for i in authors]
                return jsonify(
                    {
                        "success": True,
                        "message": f"Fetched all Research",
                        "data": [
                            {
                                "research_id": i.research_id,
                                "title": i.title,
                                "authors": authors,
                                "abstract": i.abstract,
                                "user_id": i.user_id,
                                "camp_id": i.camp_id,
                                "presented_where": i.presented_where,
                                "presentation_location": i.presentation_location,
                                "presentation_date": i.presentation_date,
                                "published_where": i.published_where,
                                "publication_date": i.publication_date,
                                "doi_or_full": i.doi_or_full,
                            }
                            for i in result
                        ],
                    }
                ), 200
        else:
            return jsonify({"success": False, "message": "No matches"}), 500

    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@bp.route("/presentation-locations", methods=["GET"])
def fetch_presentation_locations():
    session = get_session(autocommit=False, autoflush=False)()
    try:
        # Query for unique presentation locations
        presentation_locations = (
            session.query(Research.presentation_location).distinct().all()
        )
        # Generate unique IDs for each presentation location
        locations_list = [
            {"presentationLocationId": idx + 1, "presentationLocation": loc[0]}
            for idx, loc in enumerate(presentation_locations)
        ]
        return jsonify(
            {
                "success": True,
                "message": "Fetched all Presentation Locations",
                "data": locations_list,
            }
        ), 200
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@bp.route("/publication-locations", methods=["GET"])
def fetch_publication_locations():
    session = get_session(autocommit=False, autoflush=False)()
    try:
        # Query for unique publication locations
        publication_locations = session.query(Research.published_where).distinct().all()
        # Generate unique IDs for each publication location
        locations_list = [
            {"publicationLocationId": idx + 1, "publicationLocation": loc[0]}
            for idx, loc in enumerate(publication_locations)
        ]
        return jsonify(
            {
                "success": True,
                "message": "Fetched all Publication Locations",
                "data": locations_list,
            }
        ), 200
    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# from flask import render_template_string


def generate_html_table(
    data, columns_to_include, column_headers, header="", subheader=""
):
    headers = [column_headers[col] for col in columns_to_include]
    rows = [[item.get(col, "") for col in columns_to_include] for item in data]

    html_content = """
<html>
<head>
    <style>
        table {
            width: 100%;
            border-collapse: collapse;
        }
        th, td {
            border: 1px solid black;
            padding: 8px;
            text-align: center;
            font-size: 10px;
            word-wrap: break-word;
        }
        th {
            background-color: grey;
            color: whitesmoke;
        }
        .header {
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 5px;
        }
        .subheader {
            font-size: 14px;
            font-weight: bold;
            margin-bottom: 10px;
        }
    </style>
</head>
<body>
    <h2 class="header">{{ header }}</h2>  <!-- Main header -->
    <h3 class="subheader">{{ subheader }}</h3>  <!-- Subheader on a new line -->
    <table>
        <thead>
            <tr>{% for header in headers %}<th>{{ header }}</th>{% endfor %}</tr>
        </thead>
        <tbody>
            {% for row in rows %}
            <tr>{% for cell in row %}
                <td>{{ cell }}</td>
            {% endfor %}</tr>
            {% endfor %}
        </tbody>
    </table>
</body>
</html>



    """
    return render_template_string(
        html_content, headers=headers, rows=rows, header=header, subheader=subheader
    )


@bp.route("/generateReportPDF", methods=["GET"])
def generate_report_pdf():
    session = get_session(autocommit=False, autoflush=False)()

    try:
        search_query = request.args.get("q")
        startDate = request.args.get("startDate")
        endDate = request.args.get("endDate")
        presentation_location = request.args.get("presentation")
        publication_location = request.args.get("publication")
        department = request.args.get("department")
        camp_name = request.args.get("camp_name")
        columns = request.args.get("columns")
        header = request.args.get("header", "Research Report")
        subheader = request.args.get("subheader", "")

        valid_columns = {
            "research_id": "Research ID",
            "title": "Title",
            "authors": "Authors",
            "abstract": "Abstract",
            "user_id": "User ID",
            "dept_id": "Dept ID",
            "dept_name": "Department Name",
            "inst_agendas": "Institutional Agenda",
            "dept_agenda": "Departmental Agenda",
            "camp_id": "Campus ID",
            "camp_name": "Campus",
            "presented_where": "Presented Location",
            "presentation_location": "Presentation Location",
            "presentation_date": "Presentation Date",
            "published_where": "Published Where",
            "publication_date": "Publication Date",
            "doi_or_full": "DOI or Full Text",
            "cited_date": "Citation Date",
        }

        column_conversion = {
            "Author/s": "authors",
            "Title": "title",
            "Abstract": "abstract",
            "Institutional Agenda": "inst_agendas",
            "Department Agenda": "dept_agenda",
            "Presentation Location": "presentation_location",
            "Presentation Date": "presentation_date",
            "Publication Date": "publication_date",
            "Department Name": "dept_name",
            "Campus": "camp_name",
            "Citation Date": "cited_date",
        }

        if not columns:
            return jsonify(
                {"success": False, "message": "No valid columns specified"}
            ), 400

        columns_to_include = columns.split(",")
        columns_to_include = [
            column_conversion.get(col.strip(), col.strip())
            for col in columns_to_include
        ]
        columns_to_include = [col for col in columns_to_include if col in valid_columns]

        if not columns_to_include:
            return jsonify(
                {"success": False, "message": "No valid columns specified"}
            ), 400

        query = (
            session.query(Research, Departments, Campus, InstAgenda, DeptAgenda)
            # .distinct()
            .select_from(Research)
            # .group_by(Research.research_id)
            .join(
                ResearchDepartment,
                Research.research_id == ResearchDepartment.research_id,
            )
            .join(Departments, ResearchDepartment.dept_id == Departments.dept_id)
            .join(Campus, Research.camp_id == Campus.camp_id)
            .outerjoin(
                ResearchInstAgenda,
                Research.research_id == ResearchInstAgenda.research_id,
            )
            .outerjoin(
                InstAgenda, ResearchInstAgenda.instagenda_id == InstAgenda.instagenda_id
            )
            .outerjoin(
                ResearchDeptAgenda,
                Research.research_id == ResearchDeptAgenda.research_id,
            )
            .outerjoin(
                DeptAgenda, ResearchDeptAgenda.deptagenda_id == DeptAgenda.deptagenda_id
            )
            .outerjoin(
                AuthorResearch, Research.research_id == AuthorResearch.research_id
            )
            .outerjoin(Author, AuthorResearch.author_id == Author.author_id)
        )

        if department:
            departments_list = department.split(",")
            department_filters = [
                Departments.dept_name.ilike(f"%{dept.strip()}%")
                for dept in departments_list
            ]
            query = query.filter(or_(*department_filters))

        if search_query:
            search_query = f"%{search_query}%"
            query = query.filter(
                or_(
                    Author.author_name.ilike(search_query),
                    Research.title.ilike(search_query),
                    Research.presented_where.ilike(search_query),
                    InstAgenda.instagenda_name.ilike(search_query),
                    DeptAgenda.deptagenda_name.ilike(search_query),
                    Research.abstract.ilike(search_query),
                )
            )

        if presentation_location:
            presentation_list = presentation_location.split(",")
            presentation_filters = [
                Research.presentation_location.ilike(f"%{location.strip()}%")
                for location in presentation_list
            ]
            query = query.filter(or_(*presentation_filters))

        if publication_location:
            publication_list = publication_location.split(",")
            publication_filters = [
                Research.published_where.ilike(f"%{location.strip()}%")
                for location in publication_list
            ]
            query = query.filter(or_(*publication_filters))

        # Convert string dates to datetime objects if provided
        if startDate or endDate:
            if startDate:
                startDate = datetime.strptime(startDate, "%Y-%m-%d")
            if endDate:
                endDate = datetime.strptime(endDate, "%Y-%m-%d")

            # Apply filters based on the provided dates for cited_date, publication_date, and presentation_date
            if startDate and endDate:
                query = query.filter(
                    or_(
                        and_(
                            Research.cited_date >= startDate,
                            Research.cited_date <= endDate,
                        ),
                        and_(
                            Research.publication_date >= startDate,
                            Research.publication_date <= endDate,
                        ),
                        and_(
                            Research.presentation_date >= startDate,
                            Research.presentation_date <= endDate,
                        ),
                    )
                )
            elif startDate:
                query = query.filter(
                    or_(
                        Research.cited_date >= startDate,
                        Research.publication_date >= startDate,
                        Research.presentation_date >= startDate,
                    )
                )
            elif endDate:
                query = query.filter(
                    or_(
                        Research.cited_date <= endDate,
                        Research.publication_date <= endDate,
                        Research.presentation_date <= endDate,
                    )
                )

        if camp_name:
            camp_list = camp_name.split(",")
            camp_filters = [
                Campus.camp_name.ilike(f"%{camp.strip()}%") for camp in camp_list
            ]
            query = query.filter(or_(*camp_filters))

        result = query.all()

        if result:
            data = []
            for research, department, campus, instagenda, deptagenda in result:
                authors = [
                    author.author_name
                    for author in session.query(Author)
                    .join(AuthorResearch, Author.author_id == AuthorResearch.author_id)
                    .filter(AuthorResearch.research_id == research.research_id)
                    .all()
                ]
                row = {
                    "research_id": research.research_id,
                    "title": research.title,
                    "authors": ", ".join(authors),
                    "abstract": research.abstract,
                    "user_id": research.user_id,
                    "dept_id": department.dept_id,
                    "dept_name": department.dept_name,
                    "inst_agendas": instagenda.instagenda_name,
                    "dept_agenda": deptagenda.deptagenda_name,
                    "camp_id": campus.camp_id,
                    "camp_name": campus.camp_name,
                    "presented_where": research.presented_where,
                    "presentation_location": research.presentation_location,
                    "presentation_date": research.presentation_date,
                    "published_where": research.published_where,
                    "publication_date": research.publication_date,
                    "doi_or_full": research.doi_or_full,
                    "cited_date": research.cited_date,
                }
                data.append(row)

            html_content = generate_html_table(
                data, columns_to_include, valid_columns, header, subheader
            )

            pdf = pdfkit.from_string(html_content, False)
            response = make_response(pdf)
            response.headers["Content-Type"] = "application/pdf"

            return response

        else:
            return jsonify(
                {"success": True, "message": "No research found", "data": []}
            ), 200

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# Define the date formatting function
def format_date(date_obj):
    if isinstance(date_obj, datetime):
        return date_obj.strftime("%B %d, %Y")
    return ""


@bp.route("/generateReportExcel", methods=["GET"])
def generate_report_excel():
    session = get_session(autocommit=False, autoflush=False)()

    try:
        search_query = request.args.get("q")
        startDate = request.args.get("startDate")
        endDate = request.args.get("endDate")
        presentation_location = request.args.get("presentation")
        publication_location = request.args.get("publication")
        department = request.args.get("department")
        camp_name = request.args.get("camp_name")
        columns = request.args.get("columns")
        header = request.args.get("header", "")
        subheader = request.args.get("subheader", "")

        valid_columns = {
            "r  esearch_id": "Research ID",
            "title": "Title",
            "abstract": "Abstract",
            "authors": "Authors",
            "user_id": "User ID",
            "dept_name": "Department Name",
            "inst_agendas": "Institutional Agenda",
            "dept_agenda": "Departmental Agenda",
            "camp_name": "Campus",
            "presented_where": "Presented Location",
            "presentation_location": "Presentation Location",
            "presentation_date": "Presentation Date",
            "published_where": "Published Where",
            "publication_date": "Publication Date",
            "doi_or_full": "DOI or Full Text",
            "cited_date": "Citation Date",
        }

        column_conversion = {
            "Author/s": "authors",
            "Title": "title",
            "Abstract": "abstract",
            "Institutional Agenda": "inst_agendas",
            "Department Agenda": "dept_agenda",
            "Presentation Location": "presentation_location",
            "Presentation Date": "presentation_date",
            "Publication Date": "publication_date",
            "Department Name": "dept_name",
            "Campus": "camp_name",
            "Citation Date": "cited_date",
        }

        if not columns:
            return jsonify(
                {"success": False, "message": "No valid columns specified"}
            ), 400

        columns_to_include = columns.split(",")
        columns_to_include = [
            column_conversion.get(col.strip(), col.strip())
            for col in columns_to_include
        ]
        columns_to_include = [col for col in columns_to_include if col in valid_columns]

        if not columns_to_include:
            return jsonify(
                {"success": False, "message": "No valid columns specified"}
            ), 400

        query = (
            session.query(
                Research,
                ResearchDepartment,
                Departments,
                Campus,
                InstAgenda,
                DeptAgenda,
            )
            .join(
                ResearchDepartment,
                Research.research_id == ResearchDepartment.research_id,
            )
            .join(Departments, ResearchDepartment.dept_id == Departments.dept_id)
            .join(Campus, Research.camp_id == Campus.camp_id)
            .join(
                ResearchInstAgenda,
                Research.research_id == ResearchInstAgenda.research_id,
            )
            .join(
                InstAgenda, ResearchInstAgenda.instagenda_id == InstAgenda.instagenda_id
            )
            .join(
                ResearchDeptAgenda,
                Research.research_id == ResearchDeptAgenda.research_id,
            )
            .join(
                DeptAgenda, ResearchDeptAgenda.deptagenda_id == DeptAgenda.deptagenda_id
            )
        )

        if department:
            departments_list = department.split(",")
            department_filters = [
                Departments.dept_name.ilike(f"%{dept.strip()}%")
                for dept in departments_list
            ]
            query = query.filter(or_(*department_filters))

        if search_query:
            search_query = f"%{search_query}%"
            query = query.filter(
                or_(
                    Author.author_name.ilike(search_query),
                    Research.title.ilike(search_query),
                    Research.presented_where.ilike(search_query),
                    InstAgenda.instagenda_name.ilike(search_query),
                    DeptAgenda.deptagenda_name.ilike(search_query),
                    Research.abstract.ilike(search_query),
                )
            )
        if presentation_location:
            presentation_list = presentation_location.split(",")
            presentation_filters = [
                Research.presentation_location.ilike(f"%{location.strip()}%")
                for location in presentation_list
            ]
            query = query.filter(or_(*presentation_filters))

        if publication_location:
            publication_list = publication_location.split(",")
            publication_filters = [
                Research.published_where.ilike(f"%{location.strip()}%")
                for location in publication_list
            ]
            query = query.filter(or_(*publication_filters))

        # Convert string dates to datetime objects if provided
        if startDate or endDate:
            if startDate:
                startDate = datetime.strptime(startDate, "%Y-%m-%d")
            if endDate:
                endDate = datetime.strptime(endDate, "%Y-%m-%d")

            # Apply filters based on the provided dates for cited_date, publication_date, and presentation_date
            if startDate and endDate:
                query = query.filter(
                    or_(
                        and_(
                            Research.cited_date >= startDate,
                            Research.cited_date <= endDate,
                        ),
                        and_(
                            Research.publication_date >= startDate,
                            Research.publication_date <= endDate,
                        ),
                        and_(
                            Research.presentation_date >= startDate,
                            Research.presentation_date <= endDate,
                        ),
                    )
                )
            elif startDate:
                query = query.filter(
                    or_(
                        Research.cited_date >= startDate,
                        Research.publication_date >= startDate,
                        Research.presentation_date >= startDate,
                    )
                )
            elif endDate:
                query = query.filter(
                    or_(
                        Research.cited_date <= endDate,
                        Research.publication_date <= endDate,
                        Research.presentation_date <= endDate,
                    )
                )

        if camp_name:
            camp_list = camp_name.split(",")
            camp_filters = [
                Campus.camp_name.ilike(f"%{camp.strip()}%") for camp in camp_list
            ]
            query = query.filter(or_(*camp_filters))

        result = query.all()

        if result:
            data = []
            for (
                research,
                research_department,
                department,
                campus,
                instagenda,
                deptagenda,
            ) in result:
                authors = [
                    author.author_name
                    for author in session.query(Author)
                    .join(AuthorResearch, Author.author_id == AuthorResearch.author_id)
                    .filter(AuthorResearch.research_id == research.research_id)
                    .all()
                ]
                row = {
                    "research_id": research.research_id,
                    "title": research.title,
                    "abstract": research.abstract,
                    "authors": ", ".join(authors),
                    "user_id": research.user_id,
                    "dept_name": department.dept_name,
                    "inst_agendas": instagenda.instagenda_name,
                    "dept_agenda": deptagenda.deptagenda_name,
                    "camp_name": campus.camp_name,
                    "presented_where": research.presented_where,
                    "presentation_location": research.presentation_location,
                    "presentation_date": research.presentation_date,
                    "published_where": research.published_where,
                    "publication_date": research.publication_date,
                    "doi_or_full": research.doi_or_full,
                    "cited_date": research.cited_date,
                }
                data.append(row)

            # Create a DataFrame with only the selected columns
            df = pd.DataFrame(data)
            df = df[columns_to_include]

            # Generate Excel file in-memory
            output = BytesIO()

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Research Report"

            # Set font style for header, subheader, and column titles
            header_font = Font(bold=True)
            subheader_font = Font(bold=True)
            column_title_font = Font(bold=True)

            # Insert header and subheader if specified
            if header:
                ws["A1"] = header
                ws["A1"].font = header_font
            if subheader:
                ws["A2"] = subheader
                ws["A2"].font = subheader_font

            # Write valid column titles below header and subheader
            if header:
                start_row = 3  # Start after header and subheader
            else:
                start_row = 1  # Start at the beginning if no header

            for col_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(
                    row=start_row,
                    column=col_idx,
                    value=valid_columns.get(col_name, col_name),
                ).font = column_title_font

            # Insert data into worksheet
            for row_idx, row_data in enumerate(data, start=start_row + 1):
                for col_idx, col_name in enumerate(df.columns, start=1):
                    ws.cell(
                        row=row_idx, column=col_idx, value=row_data.get(col_name, "")
                    )

            for column in ws.columns:
                max_length = 0
                column = column[0].column_letter  # Get the column name
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # Save workbook to BytesIO buffer
            wb.save(output)

            output.seek(0)

            # Prepare Excel response
            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = (
                "attachment; filename=research_report.xlsx"
            )
            response.headers["Content-type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            return response

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        session.close()
